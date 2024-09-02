import tkinter as tk
from tkinter import ttk
import os
import sys
import time
from tkinter import *
import tkinter.font as font
import base64
os.environ['CUDA_VISIBLE_DEVICES'] = '0'
import pyautogui as pg
import numpy as np
import cv2
from PIL import ImageTk, Image
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from easygui import *
import easygui
import win32api,win32con



def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.getcwd()
    return os.path.join(base_path, relative_path)


i=int()
i=0
j=int()
j=1
k=int()
k=1
max_line = 21
result_list=[]
time_value=time.strftime("%d-%m-%y")
print(time_value)
output = ""
defect_name1,defect_number1,defect_letter1,defect_FAR1,SOUNDAGE1= "","","","",""

         
def save_file():
    global getname
    pg.moveTo(100,600)
    pg.click()
    pg.keyDown('ctrl')
    pg.press('s')# SAVE
    pg.keyUp('ctrl')
    #time.sleep(int(delay_save.get()))
    time.sleep(int("2"))
    #pg.(100,100)

    if var5.get()==1:
        try:
            #the_x,the_y= pg.locateCenterOnScreen('name.png', grayscale=True,confidence = confidenceE.get())
            x, y = pg.locateCenterOnScreen('name.jpg', grayscale=True,confidence =.9)
            pg.moveTo(x +200, y)
            pg.click()
            #print(x +200, y)
        except Exception as e:
            print("the exception is", e)
    if var5.get()==0:
        pg.moveTo(330,855)
        pg.click()
        #print("save coordinate= 330,855")
    time.sleep(1)
    file1 = "{}\{}".format(newpath,getname+"-{}".format(output))
    
    caps_status = win32api.GetKeyState(win32con.VK_CAPITAL)
    if caps_status==0:
        #print('CapsLock is off')
        file = file1[2:]
        pg.press('capslock')
        pg.write(str(file1[:2]))# SAVE
        pg.press('capslock')
        pg.write(file)
        #print("SAVED file======== ", file)
        pg.press('enter')
    else:
        #print('CapsLock is on')
        file = file1[2:]
        pg.write(str(file1[:2]))# SAVE
        pg.press('capslock')
        pg.write(file)
        #print("SAVED file======== ", file)
        pg.press('enter')
    
def tube_finished(event):
    global i,j,k,result_list,ws,operators_names1,getname
    
    if not os.path.isfile(r"{}\RAPPORT RX1 N°{}_{}_{}_{}.xlsx".format(path,k, EQUIPE1,POST1,time_value)):
        wb = load_workbook('RX1_template.xlsx')
        wb.save(r"{}\RAPPORT RX1 N°{}_{}_{}_{}.xlsx".format(path,k, EQUIPE1,POST1,time_value))
        
    wb = load_workbook(r"{}\RAPPORT RX1 N°{}_{}_{}_{}.xlsx".format(path,k, EQUIPE1,POST1,time_value))    
    ws = wb.active
    ws['E1'] = f"Page: {k}"
    ws['E1'].font = Font(size=18)
    #ws['D2'] = f'Rapport de contrôle\n RX2 NUMERIQUE N° "{k}"'
    # set the "Equipe" and the "Post" from the form
    #EQUIPE1 POST1 operators2_names1 operators_names1
    ws['A4'] =f'Projet:   CEEG/KD-AL.\nEquipe:   "{EQUIPE1}".\nPost:       "{POST1}".'
    # set the name of the operators
    
    
    ws['E3'] = f'DATE: {time.strftime("%d-%m-%y")}'
    ws['E3'].font = Font(size=18)
    ws['A{}'.format(j+9)] = j
    ws['B{}'.format(j+9)] = getname
    ws['C{}'.format(j+9)] = i
    
    xl_results0 =  (str(result_list)).replace('[','')
    xl_results1 =  xl_results0.replace(']','')
    xl_results2 =  xl_results1.replace("'","")
    # specifiying the width of the celule
    x = int(len(xl_results2)/49)
    #print("xl_results2=",xl_results2)
    #print("x=",x)
    if x==0:
        ws.row_dimensions[j+9].height = 20
        ws['D{}'.format(j+9)] =xl_results2
    else:
        ws.row_dimensions[j+9].height = (x+1)*17
        ws['D{}'.format(j+9)] =xl_results2
    
    ws['A31'] =f'Nom et Prénom (OP1):\n{operators_names1}\n Visa:'
    ws['D31'] =f'                                                     Nom et Prénom (OP2):\n                                                     {operators2_names1}\n                                                     Visa:'
                                             
    wb.save(r"{}\RAPPORT RX1 N°{}_{}_{}_{}.xlsx".format(path,k, EQUIPE1,POST1,time_value))
    i=0
    result_list=[]
    j+=1
    DEFECT_NUMB_lab.config(text=i)
    finish_tube.config(text="Tube N°{}".format(j),bg ="RoyalBlue1")
    
    if j== max_line and var2.get() == 1:
        win32api.ShellExecute(
        0,
        "print",
        r"{}\RAPPORT RX1 N°{}_{}_{}_{}.xlsx".format(path,k, EQUIPE1,POST1,time_value),
        None,
        ".",
        0
        )
        #os.startfile(r"{}\POST REPORT N°{} at {}.xlsx".format(path,k,time_value), 'print')
        report_closed.config(text="P{}-PRINTED".format(k),bg = "green2")
        finish_tube.config(text="INSÈRE",bg =btncolor)
        k+=1
        j=1
        result_list=[]
        
    #rest all widget
    PIPE_NAME.delete(0,"end")
    defect_name.set("")
    defect_number.set("")
    defect_letter.set("")
    defect_FAR.set("") 
    print("!!! tube finished !!!")
    #devlabel.config(text= "Tube finished!",fg="orange",bg="yellow")
    
    PIPE_NAME.focus_set()
    
    
def report_closed_func(event):
    global k,i,j
    if var2.get() == 1:   

        # PRINT THE REPPORT:
        win32api.ShellExecute(
        0,
        "print",
        r"{}\RAPPORT RX1 N°{}_{}_{}_{}.xlsx".format(path,k, EQUIPE1,POST1,time_value),
        None,
        ".",
        0
        )
        #rest all widget
        #os.startfile(r"{}\POST REPORT N°{} at {}.xlsx".format(path,k,time_value), 'print')
        PIPE_NAME.delete(0,"end")
        defect_name.set("")
        defect_number.set("")
        defect_letter.set("")
        defect_FAR.set("")
        print("!!!!!!!!rapport closed!!!!!!!!")
        devlabel.config(text= "Rapport closed!",fg="orange",bg="yellow")
        
        report_closed.config(text="P{}-PRINTED".format(k),bg = "green2")
        finish_tube.config(text="INSÈRE",bg =btncolor)
        k+=1
        i=0
        j=1
        PIPE_NAME.focus_set()


def open_and_create_folder(event):
    
    global i,newpath,path,wb,result_list,output,defect_name1,defect_number1,defect_letter1,defect_FAR1,getname
    pipe_name=str(PIPE_NAME.get())

    if PIPE_NAME.get()=="" or pathE.get()=="":
        print("File up pipe name, path entries and defect name ")
        devlabel.config(text= "File up pipe name,\n path entries and defect name!!",fg="red",bg="yellow")
        return
    if len(pipe_name)==2:
        print("length====",len(pipe_name),pipe_name[1])
        #A0001
        getname = (pipe_name[0]+"0"+"0"+"0"+pipe_name[1]).upper()
        print("getname",getname)
    elif len(PIPE_NAME.get())==3:
        #A0012
        getname = (pipe_name[0]+"0"+"0"+pipe_name[1]+pipe_name[2]).upper()
    elif len(PIPE_NAME.get())==4:
        #A0123
        getname = (pipe_name[0]+"0"+pipe_name[1]+pipe_name[2]+pipe_name[3]).upper()
    elif len(PIPE_NAME.get())==5:
        #A1234
        devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-INSPECTOR RT-II 2020",fg="black",bg=color)
        getname = str(PIPE_NAME.get()).upper()
        print("getname= ",getname)
    elif ((PIPE_NAME.get()).upper().find('-BIS') != -1):
        devlabel.config(text= "BIS TUBE!!!",fg="black",bg="yellow")
        getname = str(PIPE_NAME.get()).upper()
    name  = str(getname[0])
    path = pathE.get()
    newpath = "{}\{}\{}".format(path,name,getname)
            
    if i==0 and os.path.exists(newpath):
        print("file exists!!!XXXXXXXXXXXXX!!")
        devlabel.config(text= "Folder exists!!!",fg="orange",bg="yellow")
        
    
    if not os.path.exists(newpath):
        os.makedirs(newpath)
        
    os.startfile(newpath)
    PIPE_NAME.delete(0,"end")
    PIPE_NAME.insert(0,getname)

        
    
def integration(event):
    global i,newpath,path,wb,result_list,output,defect_name1,defect_number1,defect_letter1,defect_FAR1,SOUNDAGE1,getname
    pipe_name=str(PIPE_NAME.get())

    if PIPE_NAME.get()=="" or pathE.get()==""or defect_name1=="":
        print("File up pipe name, path entries and defect name ")
        devlabel.config(text= "File up pipe name,\n path entries and defect name!!",fg="red",bg="yellow")
        return
    if len(pipe_name)==2:
        print("length====",len(pipe_name),pipe_name[1])
        #A0001
        getname = (pipe_name[0]+"0"+"0"+"0"+pipe_name[1]).upper()
        print("getname",getname)
    elif len(PIPE_NAME.get())==3:
        #A0012
        getname = (pipe_name[0]+"0"+"0"+pipe_name[1]+pipe_name[2]).upper()
    elif len(PIPE_NAME.get())==4:
        #A0123
        getname = (pipe_name[0]+"0"+pipe_name[1]+pipe_name[2]+pipe_name[3]).upper()
    elif len(PIPE_NAME.get())==5:
        #A1234
        devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-INSPECTOR RT-II 2020",fg="black",bg=color)
        getname = str(PIPE_NAME.get()).upper()
        print("getname= ",getname)
    elif ((PIPE_NAME.get()).upper().find('-BIS') != -1):
        devlabel.config(text= "BIS TUBE!!!",fg="black",bg="yellow")
        getname = str(PIPE_NAME.get()).upper()
        
    else:
        devlabel.config(text= "UNCORRECT TUBE NUMBER!",fg="black",bg=color)
        return
    name  = str(getname[0])
    path = pathE.get()
    newpath = "{}\{}\{}".format(path,name,getname)
            
    if i==0 and os.path.exists(newpath):
        print("file exists!!!XXXXXXXXXXXXX!!")
        devlabel.config(text= "Folder exists!!!",fg="orange",bg="yellow")
    if not os.path.exists(newpath):
        os.makedirs(newpath)
    i+=1
    time_integration = int("9")
    print("Integration started!")
    print("i=",i)

    #time.sleep(time_integration)
    pg.press('f4')# INTEGRATE
    pg.moveTo(167,55)
    pg.click()# CONFIRM INTEGRATION
    print("defect number=",defect_name1)
    
    time.sleep(8)
    output = "{}{}{}{}{}".format(defect_name1,defect_number1,defect_letter1,SOUNDAGE1,defect_FAR1)
    """
    if var2.get() == 1:
        output = "{}{}{}{}RM1{}".format(defect_name1,defect_number1,defect_letter1,SOUNDAGE1,defect_FAR1)#f'{defect_name.get()}{defect_number.get()}{defect_letter.get()}RM1{defect_FAR.get()}'
        print("output=",output)
    else:
        output = "{}{}{}{}{}".format(defect_name1,defect_number1,defect_letter1,SOUNDAGE1,defect_FAR1)"""
    if output:
        print("output=",output)
        result_list.append(output)
        
        defect_name1,defect_number1,SOUNDAGE1,defect_letter1,defect_FAR1= "","","","",""
        defect_name.set("")
        SOUNDAGE.set("")
        defect_number.set("")
        defect_letter.set("")
        defect_FAR.set("")
        # deselect the checkbox rm1
        RM1.deselect()

    save_file()
  
    pg.moveTo(120,55)# PRESS LIVE
    pg.click()
    pg.moveTo(120,55)# PRESS LIVE
    pg.click()
    #time.sleep(int(DELAYWORKER.get()))

    DEFECT_NUMB_lab.config(text = " {}".format(i), bg ="yellow")
##        time_delay = int("0")
##        print("time delay = ",time_delay)
##        time.sleep(time_delay)
    print('finished')
    defect_name.focus_set()
    devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-INSPECTOR RT-II 2020",fg="black",bg=color)
    PIPE_NAME.delete(0,"end")
    PIPE_NAME.insert(0,getname)


def check5(event):
    global j
    if var7.get()!=1:
        j_variable_lab.grid(row = 7 , column= 0, padx = 10,sticky="W")
        j_variable.grid(row = 7 , column= 1, padx = 10,sticky="W")
        j_variable_btn.grid(row =7,column=1, padx = 15,pady = 10,sticky="E")
        k_variable_lab.grid(row = 8 , column= 0, padx = 10,sticky="W")
        k_variable.grid(row = 8 , column= 1, padx = 10,sticky="W")
        #j = int(j_variable.get())
        print("var7=", var7.get(),"j=",j)
        root.geometry("410x395")
                
    else:
        j_variable.grid_forget()
        j_variable_lab.grid_forget()
        j_variable_btn.grid_forget()
        k_variable_lab.grid_forget()
        k_variable_btn.grid_forget()
        
        print("var7=", var7.get(),"j=",j)
        root.geometry("410x395")


def j_variablefunc():
    global j,k
    if var7.get()==1:
        j = int(j_variable.get())
        k = int(k_variable.get())
        j_variable_btn.config(text=f"R N°{k}/L N°{j}",bg="green2")
        print("var7=", var7.get(),"j,k=",j,k)

    
def quitt():
    root.quit()
    sys.exit()
    #top = Toplevel()

def show_frame(frame):
    frame.tkraise()
    frame.grid(row =0,column=0,sticky='nsew')




root = Tk()
root.config(bg ="white")
root.focus_force()
root.rowconfigure(0,weight= 1)
root.columnconfigure(0,weight=1)

color = "light sky blue"
btncolor ="gold"



starting_FRAME = tk.Frame(root, width=100, height=100, background=color)
starting_FRAME.grid(row =0,column=0,sticky='nsew')

PIPE_FRAME_FRAME = tk.Frame(root, width=50, height=50, background=color )
#PIPE_FRAME_FRAME.pack(fill= "both", expand ="YES" , pady = 0, padx=10)
#PIPE_FRAME_FRAME.config(width=500, bg ="white")

starting = LabelFrame(starting_FRAME, text = "INFO",  width = 35,height =20,font =("Helvetica",10,"bold"),bg =color )
starting.grid(row =0,column=0,padx=2)

#PIPE_FRAME = LabelFrame(PIPE_FRAME_FRAME, text = "INFO",  width = 35,height =20,font =("Helvetica",10,"bold"),bg =color )
#PIPE_FRAME.grid(row =0,column=0,sticky='nsew')

PIPE_ET_DESIGNATION = LabelFrame(PIPE_FRAME_FRAME, text = "PIPE ET DESIGNATION:",  width = 35,height =20,font =("Helvetica",10,"bold"),bg =color,labelanchor = "n")
PIPE_ET_DESIGNATION.grid(row =0,column=0,sticky='nsew', columnspan=4,padx= 10,pady = 10)

#ACTIONS_FRAME = LabelFrame(PIPE_FRAME_FRAME, text = "ACTIONS:",  width = 35,height =20,font =("Helvetica",10,"bold"),bg =color )
#ACTIONS_FRAME.grid(row =2,column=0,sticky='nsew', columnspan=4)


def focus_func(event):
    global operators_names1
    if str(operators_names.get())!="":
        operators2_names.focus_set()
        operator1_names_lab.config(bg = "green2")
        operators_names1  = operators_names.get()
        print(operators_names1)

    else:
        operator1_names_lab.config(bg = "orange1")



    
################################################################################################ STARTING FRAME ##########################################################################################

#OPERATOR1 LABEL

operator1_names_lab = Label(starting, text= "OPERATEUR 01:",font =("Helvetica",10,"bold"), bg =color)
operator1_names_lab.grid(row = 0 , column= 0, padx = 10,sticky="W", pady=5)

"""

# combobox5
def operators_namesfunc(event):
    global operators_names1
    operators_names1  = operators_names.get()
    print(operators_names1)

def operators_namesdel(event):
    global operators_names1
    operators_names1  = ""
    print("operators_names1 is deleted!")

## Adding combobox FILM A REFAIR
n4 = tk.StringVar() 
 
operators_names = ttk.Combobox(starting, width = 15, textvariable = n4,font= ("Courier", 16, "bold") )

operators_names['values'] =('ZAI ABDELWAHAB',  
                             'LASSOAD HAKIM', 
                             'GUERBATI ABDELOUAHAB',                            
                             'LARBI CHIKH',
                             'MEGHAZI OUSSAMA',
                             'ZERBANI MOHAMMED',
                             'LAOUFI MAAMMAR',
                             'YELLAOUI ADEL')
  
operators_names.grid(column = 1, row = 0,padx = 10, pady=5) 
operators_names.current()
operators_names.bind("<<ComboboxSelected>>", operators_namesfunc)
operators_names.bind("<BackSpace>", operators_namesdel)
operators_names.focus_set()
#########"""
operators_names = Entry(starting, width = 18,relief ="sunken", font =("Helvetica",16), bg ="white")
operators_names.grid(column = 1, row = 0,padx = 10, pady=5)
operators_names.bind("<Return>", focus_func)


# combobox6


#OPERATOR2 LABEL

operator2_names_lab = Label(starting, text= "OPERATEUR 02:",font =("Helvetica",10,"bold"), bg =color)
operator2_names_lab.grid(row = 1 , column= 0, padx = 10,sticky="W", pady=5)

def focus_func1(event):
    global operators2_names1
    if str(operators2_names.get())!="":
        POST.focus_set()
        operators2_names1  = operators2_names.get()
        operator2_names_lab.config(bg = "green2")
        print(operators2_names1)
    else:
        operator2_names_lab.config(bg = "orange1")
        
        
    
"""
# combobox5
def operators2_namesfunc(event):
    global operators2_names1
    operators2_names1  = operators2_names.get()
    print(operators2_names1)

def operators2_namesdel(event):
    global operators2_names1
    operators2_names1  = ""
    print("operators2_names1 is deleted!")

## Adding combobox FILM A REFAIR
n4 = tk.StringVar() 
 
operators2_names = ttk.Combobox(starting, width = 15, textvariable = n4,font= ("Courier", 16, "bold") )

operators2_names['values'] =('ZAI ABDELWAHAB',  
                             'LASSOAD HAKIM', 
                             'GUERBATI ABDELOUAHAB',                            
                             'LARBI CHIKH',
                             'MEGHAZI OUSSAMA',
                             'ZERBANI MOHAMMED',
                             'LAOUFI MAAMMAR',
                             'YELLAOUI ADEL')
  
operators2_names.grid(column = 1, row = 1,padx = 10, pady=5) 
operators2_names.current()
operators2_names.bind("<<ComboboxSelected>>", operators2_namesfunc)
operators2_names.bind("<BackSpace>", operators2_namesdel)
"""
#############
operators2_names = Entry(starting, width = 18,relief ="sunken", font =("Helvetica",16), bg ="white")
operators2_names.grid(column = 1, row = 1,padx = 10, pady=5)
operators2_names.bind("<Return>", focus_func1)
#PIPE_NAME.focus_set()
#PIPE_NAME.bind("<Return>", open_and_create_folder)


# POST LABEL

POST_lab = Label(starting, text= "POSTE:",font =("Helvetica",10,"bold"), bg =color)
POST_lab.grid(row = 2 , column= 0, padx = 10,sticky="W", pady=5)


def POSTfunc(event):
    global POST1
    POST1  = POST.get()
    POST_lab.config(bg= "green2")
    print(POST1)

def POSTdel(event):
    global POST1
    POST1  = ""
    print("POST1 is deleted!")

n5 = tk.StringVar() 
 
POST = ttk.Combobox(starting, width = 15, textvariable = n5,font= ("Courier", 16, "bold") )

POST['values'] =('1 er',  
                 '2 eme', 
                 '3 eme')
  
POST.grid(column = 1, row = 2,padx = 10, pady=5) 
POST.current()
POST.bind("<<ComboboxSelected>>", POSTfunc)
POST.bind("<BackSpace>", POSTdel)


# EQUIPE

# EQUIPE LABEL

EQUIPE_lab = Label(starting, text= "GROUPE:",font =("Helvetica",10,"bold"), bg =color)
EQUIPE_lab.grid(row = 3 , column= 0, padx = 10,sticky="W", pady=5)


def EQUIPEfunc(event):
    global EQUIPE1
    EQUIPE1  = EQUIPE.get()
    EQUIPE_lab.config(bg = "green2")
    print(EQUIPE1)

def EQUIPEdel(event):
    global EQUIPE1
    EQUIPE1  = ""
    print("EQUIPE1 is deleted!")

n5 = tk.StringVar() 
 
EQUIPE = ttk.Combobox(starting, width = 15, textvariable = n5,font= ("Courier", 16, "bold") )

EQUIPE['values'] =('A',  
                 'B', 
                 'C',
                 'D')
  
EQUIPE.grid(column = 1, row = 3,padx = 10, pady=5) 
EQUIPE.current()
EQUIPE.bind("<<ComboboxSelected>>", EQUIPEfunc)
EQUIPE.bind("<BackSpace>", EQUIPEdel)


pathE = Entry(starting, width = 19,relief ="groove", font =("Helvetica",15),bg="white")
pathE.grid(row = 4 , column= 1, pady = 5,columnspan = 3)
pathE.insert(0,"Y:\partage\CEEG\CEEG 2021")
#pathE.insert(0,"C:\\Users\\111\\Desktop\\EGPDF")

pathE_lab = Label(starting, text= "PATH:",font =("Helvetica",10,"bold"), bg =color)
pathE_lab.grid(row = 4 , column= 0, padx = 10,sticky="W")

var7 = IntVar()

continueing_checkbtn= Checkbutton(starting, text = "RAPPORT INCOMPLET",font =("Helvetica",8,"bold"), variable = var7, bg =color)
continueing_checkbtn.grid(row = 6 , column= 0,padx = 10, pady = 5,sticky="W")
continueing_checkbtn.bind('<Button-1>',check5)


j_variable_lab = Label(starting, text= "Line N°:",font =("Helvetica",10,"bold"), bg =color)
j_variable = Spinbox(starting,from_=1, to = max_line ,bg ="white",increment =1,width =3, font =("Helvetica",13),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")
#j_variable.grid(row = 4 , column= 1, pady = 10,columnspan = 2)
#j_variable.delete(0,"end")
#j_variable.insert(0,"8")
#j_variable.bind("<<SpinboxSelected>>",j_variablefunc)

k_variable_lab = Label(starting, text= "Report N°:",font =("Helvetica",10,"bold"), bg =color)
k_variable = Spinbox(starting,from_=1, to = max_line ,bg ="white",increment =1,width =3, font =("Helvetica",13),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")



j_variable_btn= tk.Button(starting,text="VALIDE",bg = btncolor,command=j_variablefunc,font =("Helvetica",10,"bold"),height = 2, width = 10)

fr1_btn= tk.Button(starting,text="ENTER",bg = btncolor,command=lambda:show_frame(PIPE_FRAME_FRAME),font =("Helvetica",10,"bold"),height = 2, width = 10)
fr1_btn.grid(row =6,column=1, padx = 15,pady = 20,sticky ="E")

################################################################################################ END STARTING FRAME ##########################################################################################


x_space=10

############################################################################################### PIPE_FRAME ############################################################################################################



# PIPE_FRAME WEDGITS

PIPE_NAME_lab = Label(PIPE_ET_DESIGNATION, text= "PIPE: ",font =("Helvetica",15,"bold"), bg =color)
PIPE_NAME_lab.grid(row = 0 , column= 0, padx =0,sticky="W")


PIPE_NAME = Entry(PIPE_ET_DESIGNATION, width = 6,relief ="sunken", font =("Helvetica",20), bg ="white")
PIPE_NAME.grid(row = 0 , column= 1, padx = x_space, pady = 10)
PIPE_NAME.focus_set()
PIPE_NAME.bind("<Return>", open_and_create_folder)

# defect name label

#defect_name_lab = Label(PIPE_ET_DESIGNATION, text= "DESIGNATION:",font =("Helvetica",10,"bold"), bg =color)
#defect_name_lab.grid(row = 1 , column= 0, padx = 10,sticky="W")



# combobox1
font1= ("Courier", 20, "bold")

def defect_namefunc(event):
    global defect_name1
    defect_name1 = defect_name.get()
    print(defect_name1)

def defect_namedel(event):
    global defect_name1
    defect_name1 = ""
    print("defect_name1 is delated!")


n = tk.StringVar() 
defect_name = ttk.Combobox(PIPE_ET_DESIGNATION, width = 2, textvariable = n,font= font1) 
  
# Adding combobox DEFECT NAME:
defect_name['values'] =  ('U', 
                          'E', 
                          'Y', 
                          'EY', 
                          'S')
  
defect_name.grid(row = 0,column = 2, padx = x_space, pady = 10) 
defect_name.current()
defect_name.bind("<<ComboboxSelected>>", defect_namefunc)
defect_name.bind("<BackSpace>", defect_namedel)


# combobox2.5

def defect_numberfunc(event):
    global defect_number1
    defect_number1= defect_number.get()
    print(defect_number1)
    
def defect_numberdel(event):
    global defect_number1
    defect_number1= ""
    print("defect_number1 is delated!")


    
    
n1 = tk.StringVar() 
defect_number = ttk.Combobox(PIPE_ET_DESIGNATION, width = 2, textvariable = n1,font= font1) 

#defect_name,defect_number,defect_letter,defect_FAR


    
# Adding combobox NUMBER OF DEFECTS:
defect_number['values'] =('1',  
                          '2', 
                          '3', 
                          '4', 
                          '5', 
                          '6', 
                          '7',
                          '8', 
                          '9', 
                          '10')
  
defect_number.grid(row = 0,column = 3, padx = x_space, pady = 10) 
defect_number.current()
defect_number.bind("<<ComboboxSelected>>", defect_numberfunc)
defect_number.bind("<BackSpace>", defect_numberdel)



  

# combobox3

def defect_letterfunc(event):
    global defect_letter1
    defect_letter1 = defect_letter.get()
    print(defect_letter1)
    
def defect_letterdel(event):
    global defect_letter1
    defect_letter1 = ""
    print("defect_letter1 is deleted!")



    
    
# Adding combobox LETTER OF DEFECT
n2 = tk.StringVar() 
 
defect_letter = ttk.Combobox(PIPE_ET_DESIGNATION, width = 1, textvariable = n2,font= font1 )

defect_letter['values'] =('A',  
                          'B', 
                          'C', 
                          'D', 
                          'E', 
                          'F', 
                          'G',
                          'H', 
                          'I', 
                          'J')
  
defect_letter.grid(row = 0,column = 4, padx = x_space, pady = 10) 
defect_letter.current()
defect_letter.bind("<<ComboboxSelected>>", defect_letterfunc)
defect_letter.bind("<BackSpace>", defect_letterdel)



# combobox4
def defect_FARfunc(event):
    global defect_FAR1
    defect_FAR1  = defect_FAR.get()
    print(defect_FAR1)

def defect_FARdel(event):
    global defect_FAR1
    defect_FAR1  = ""
    print("defect_FAR1 is deleted!")

    
# combobox2

def SOUNDAGEfunc(event):
    global SOUNDAGE1
    SOUNDAGE1= SOUNDAGE.get()
    print(SOUNDAGE1)
    
def SOUNDAGEdel(event):
    global SOUNDAGE1
    SOUNDAGE1= ""
    print("defect_number1 is delated!")


    
    
n9 = tk.StringVar() 
SOUNDAGE = ttk.Combobox(PIPE_ET_DESIGNATION, width = 3, textvariable = n9,font= font1) 

#defect_name,defect_number,defect_letter,defect_FAR


    
# Adding combobox NUMBER OF DEFECTS:
SOUNDAGE['values'] =('SD', 
                     'SD1', 
                     'SD2', 
                     'SD3', 
                     'SG',
                     'SG1', 
                     'SG2',
                     'SG3')
  
SOUNDAGE.grid(row = 1,column = 1, padx = x_space, pady = 10) 
SOUNDAGE.current()
SOUNDAGE.bind("<<ComboboxSelected>>", SOUNDAGEfunc)
SOUNDAGE.bind("<BackSpace>", SOUNDAGEdel)


## Adding combobox FILM A REFAIR
n3 = tk.StringVar() 
 
defect_FAR = ttk.Combobox(PIPE_ET_DESIGNATION, width = 3, textvariable = n3,font= ("Courier", 20) )

defect_FAR['values'] =('.',  
                       '..', 
                       '...', 
                       '....'
                    )
  
#defect_FAR.grid(row =1,column = 2, padx = x_space, pady = 10,columnspan=2,sticky="w") 
defect_FAR.current()
defect_FAR.bind("<<ComboboxSelected>>", defect_FARfunc)
defect_FAR.bind("<BackSpace>", defect_FARdel)



#check

    

var2 = IntVar()
#var2.set(1)
v=1
def check(event):
    global v
    if v==1:
        var2.set(1)
        v=0
    elif v==0:
        var2.set(0)
        v=1
    

RM1= Checkbutton(PIPE_ET_DESIGNATION, text = "PRINT", variable = var2, bg =color)
RM1.grid(row = 1 , column= 3,padx = x_space, pady = 10,columnspan=2)
#RM1.select()
RM1.bind('<Return>',check)
RM1.bind('<Button-1>',check)

# INTEGRATION BUTTON
#######################################################################################################################################################################
###################################################################### MAIN FRAME #######################################################################################

# line 2

INTGRATIONe_lab = Label(PIPE_FRAME_FRAME, text= "INT/TUBE:",font =("Helvetica",10,"bold"), bg =color)
INTGRATIONe_lab.grid(row = 1 , column= 0, padx = 10, pady = 10,sticky="W")


DEFECT_NUMB_lab = Label(PIPE_FRAME_FRAME, text= i,font =("Helvetica",10,"bold"), bg =color)
DEFECT_NUMB_lab.grid(row = 1 , column= 1, padx = 10, pady = 10)

###################################################################### ACTIONS_FRAME: #######################################################################################


#actions_lab = Label(ACTIONS_FRAME, text= "ACTIONS:",font =("Helvetica",10,"bold"), bg =color)
#actions_lab.grid(row = 4 , column= 0, padx = 10, pady = 10,sticky="W")


# buttons
INTEGRATION_btn = Button(PIPE_FRAME_FRAME, text = "INTÉGRE", bg =btncolor,activebackground="YELLOW1",font =("Helvetica",10,"bold"),height = 2, width = 13)
INTEGRATION_btn.grid(row = 2 , column= 0,padx = 10, pady = 10,columnspan=1)
INTEGRATION_btn.bind('<Return>',integration)
INTEGRATION_btn.bind('<Button-1>',integration)

finish_tube = Button(PIPE_FRAME_FRAME, text = "INSÈRE", bg =btncolor,activebackground="orange",font =("Helvetica",10,"bold"),height = 2, width = 13)
finish_tube.grid(row = 2 , column= 1,padx = 5, pady = 10)
finish_tube.bind('<Return>',tube_finished)
finish_tube.bind('<Button-1>',tube_finished)

report_closed = Button(PIPE_FRAME_FRAME, text = "CLÔTURE", bg =btncolor,activebackground="red",font =("Helvetica",10,"bold"),height = 2, width = 13)
report_closed.grid(row = 2 , column= 2,padx = 5, pady = 10)
report_closed.bind('<Return>',report_closed_func)
report_closed.bind('<Button-1>',report_closed_func)

#QUIT_btn = Button(ACTIONS_FRAME, text = "QUIT", bg =btncolor,activebackground="RED",font =("Helvetica",10,"bold"),height = 2, width = 13, command = quitt)
#QUIT_btn.grid(row = 5 , column= 2,padx = 5, pady = 10)
#QUIT_btn.bind('<Return>',quitt)
#QUIT_btn.bind('<Button-1>',quitt)

var5 = IntVar()

use_saerch_image = Checkbutton(PIPE_FRAME_FRAME, text = "AUTO", variable = var5, bg =color)
use_saerch_image.grid(row = 1, column= 2,padx = 30, pady =10,sticky="w")
#use_saerch_image.select()




# line 3
devlabel = Label(PIPE_FRAME_FRAME, text= "DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2020",font =("Algerian",12,"bold"), bg =color)
devlabel.grid(row = 6 , column= 0, columnspan = 7, pady = 20, padx= 60)

fr3_btn= tk.Button(PIPE_FRAME_FRAME,text="RETOUR",bg = btncolor,command=lambda:show_frame(starting_FRAME),font =("Helvetica",10,"bold"),height = 2, width = 10)
fr3_btn.grid(row = 11 , column= 1,padx = 10)


############################################################################################### END PIPE_FRAME ############################################################################################################




icon = """ """


    
icondata= base64.b64decode(icon)
## The temp file is icon.ico
tempFile= "icon.ico"
iconfile= open(tempFile,"wb")
## Extract the icon
iconfile.write(icondata)
iconfile.close()
root.wm_iconbitmap(tempFile)
#top.wm_iconbitmap(tempFile)

## Delete the tempfile
os.remove(tempFile)

#top.title("CONFIGURATION")
root.title("RX2-REPORTER")
root.geometry("410x390")
root.call('wm', 'attributes', '.', '-topmost', True)
root.resizable(False,False)
root.protocol('WM_DELETE_WINDOW', quitt)
root.mainloop()
